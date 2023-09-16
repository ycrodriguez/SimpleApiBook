import openpyxl
from base.settings import BASE_DIR
from book.models import Author, Editorial, Gender, Reader


def load_authors():
    path = '{}{}'.format(BASE_DIR, '/book/scripts/files/author.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        name = sheet.cell(row=i, column=1).value
        country = sheet.cell(row=i, column=2).value
        birthdate = sheet.cell(row=i, column=3).value
        biography = sheet.cell(row=i, column=4).value
        Author.objects.get_or_create(name=name, country=country, birthdate=birthdate, biography=biography)


def load_editorial():
    path = '{}{}'.format(BASE_DIR, '/book/scripts/files/editorial.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        name = sheet.cell(row=i, column=1).value
        country = sheet.cell(row=i, column=2).value
        Editorial.objects.get_or_create(name=name, country=country)


def load_gender():
    path = '{}{}'.format(BASE_DIR, '/book/scripts/files/gender.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        name = sheet.cell(row=i, column=1).value
        Gender.objects.get_or_create(name=name)


def load_reader():
    path = '{}{}'.format(BASE_DIR, '/book/scripts/files/reader.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        name = sheet.cell(row=i, column=1).value
        address = sheet.cell(row=i, column=2).value
        phone = sheet.cell(row=i, column=3).value
        email = sheet.cell(row=i, column=4).value
        Reader.objects.get_or_create(name=name, address=address, phone=phone, email=email)


def load_all():
    load_authors()
    load_editorial()
    load_gender()
    load_reader()
